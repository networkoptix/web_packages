import csv
import datetime
import os
import zipfile
from enum import StrEnum
from io import (
    BytesIO,
    StringIO,
)
from typing import (
    Any,
    List,
    Optional,
    Union,
)

import structlog
from dateutil import parser
from dateutil.relativedelta import relativedelta
from django.conf import settings
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import (
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from partners.models import (
    ChannelPartner,
    ChannelPartnerService,
    Organization,
)
from partners.services.usage_reports_service import (
    BeginningOfPeriodDate,
    ChannelPartnerRegularUsageList,
    ChannelPartnerReportsService,
    ChannelPartnerUsageReport,
    ChannelPartnerUsageReportRecord,
    ExpiringUsageDetailRecord,
    OrganizationRegularUsageList,
    OrganizationReportsService,
    OrganizationUsageReportRecord,
    RegularUsageDetailRecord,
    TotalUsageDate,
)


logger = structlog.getLogger(__name__)


class ReportFormat(StrEnum):
    xlsx = 'xlsx'
    csv = 'csv'


class Styling:
    black = "000000"
    peach = 'FCE5CD'

    font = Font(name='Arial', size=10)
    font_bold_italic = Font(name='Arial', size=10, bold=True, italic=True)
    font_bold = Font(name='Arial', size=10, bold=True)
    font_italic = Font(name='Arial', size=10, italic=True)

    border_thin = Side(border_style="thin", color=black)
    border_dotted = Side(border_style="dotted", color=black)
    border_medium = Side(border_style="medium", color=black)
    border_double = Side(border_style="double", color=black)

    format_number = '#,##0'
    format_currency = '"$"#,##0.00'
    format_total_sum = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
    format_date = 'M/d/yyyy'

    fill_formulae = PatternFill('solid', fgColor=peach)


class HeadersMixin:

    def report_name_header(self):
        if self.service:
            return f"{self.entity.name} {self.service.name} Service Usage Report"
        else:
            return f"{self.entity.name} Service Usage Report"

    def date_range_header(self):
        return f"Date Range: {self.period_start: %B %d} - {self.period_end:%B %d, %Y}"

    def gen_date_header(self):
        return f"Report Generation Date: {self.report_date:%B %d, %Y}"


class ReportsExportServiceBase(HeadersMixin):
    """
    Base class for exporting reports. This class should not be instantiated directly.
    Instead, use one of the subclasses that implements the required methods.
    """
    generate = True
    record_row_class: Union['ExpiringServiceChangesRow', 'RegularServiceChangesRow'] = None
    columns = 0
    start_row = 6
    start_col = 2

    def __init__(
            self,
            report_date: datetime.date,
            period_start: datetime.date,
            entity: ChannelPartner | Organization,
            service: ChannelPartnerService | None,
            sheet: Worksheet
    ):
        self.report_date = report_date
        self.period_start = period_start
        self.period_end = period_start + relativedelta(months=1) - relativedelta(days=1)
        self.entity = entity
        self.service = service
        self.sheet = sheet
        self.current_row = self.start_row
        self.sub_total_rows: List[int] = []

    def generate_report(self):
        raise NotImplemented("This method must be implemented in the subclass.")

    def make_header(self):
        """
        Fill headers texts in the first 3 rows of the report.
        Preserves the template styling.
        """
        cell = self.sheet.cell(row=2, column=self.start_col)
        cell.value = self.report_name_header()
        cell = self.sheet.cell(row=3, column=self.start_col)
        cell.value = self.date_range_header()
        cell = self.sheet.cell(row=4, column=self.start_col)
        cell.value = self.gen_date_header()

    def get_report_data(self):
        """
        Get report data for the current entity (and service).
        """
        raise NotImplemented("This method must be implemented in the subclass.")

    def get_used_by(self, usages: ChannelPartnerUsageReportRecord | OrganizationUsageReportRecord):
        """
        'Used by' column for service usage report sheets.
        """
        used_by_name = (usages.get('channel_partner_name') or usages.get('organization_name')
                        or usages.get('system_name') or usages.get('system_id'))
        return used_by_name

    def fill_services(self, report: ChannelPartnerRegularUsageList | OrganizationRegularUsageList):
        """
        Fill the report data into the sheet.
        Gets data and styling from the record_row_class and applies it to a cell.
        """
        for usages in report:
            used_by_name = self.get_used_by(usages)
            for usage_record in usages['report']:
                usage_row = self.record_row_class(usage_record, self.current_row, used_by=used_by_name)
                for col_idx, (val, fmt, font, border) in enumerate(usage_row.zip_values()):
                    cell = self.sheet.cell(row=self.current_row, column=col_idx + self.start_col)
                    cell.value = val
                    cell.font = font
                    if fmt:
                        cell.number_format = fmt
                    cell.border = border
                self.current_row += 1
                used_by_name = None
            self.sub_total_rows.append(self.current_row - 1)

    def format_footer(self):
        """
        Format the footer row of the report.
        """
        for i in range(self.start_col, self.start_col + self.columns):
            cell = self.sheet.cell(row=self.current_row, column=i)
            left = Styling.border_thin
            right = Styling.border_thin
            if i == self.start_col:
                left = Styling.border_medium
            if i == self.start_col + self.columns - 1:
                right = Styling.border_medium
            cell.border = Border(top=Styling.border_double, bottom=Styling.border_medium, left=left, right=right)
            cell.number_format = Styling.format_number
            cell.font = Styling.font_bold

    def set_footer_values(self):
        """
        Set the values of the footer row of the report.
        """
        raise NotImplemented("This method must be implemented in the subclass.")

    def generate_report_sheet(self):
        """
        Generate the report sheet.
        """
        self.make_header()
        self.fill_services(self.get_report_data())
        self.format_footer()
        self.set_footer_values()


class PartnerReportDataMixin:
    def get_report_data(self):
        """
        Base data getter for a partner report.
        """
        kwargs = dict(
            channel_partner=self.entity,
            service=self.service,
            period_start=self.period_start,
            generate=self.generate
        )
        if self.service.is_expiring:
            partners_report = ChannelPartnerReportsService.get_expiring_channel_partner_usages(**kwargs)
            organizations_report = ChannelPartnerReportsService.get_expiring_organization_usages(**kwargs)
        else:
            partners_report = ChannelPartnerReportsService.get_regular_channel_partner_usages(**kwargs)
            organizations_report = ChannelPartnerReportsService.get_regular_organization_usages(**kwargs)
        return partners_report + organizations_report


class OrganizationReportDataMixin:
    def get_report_data(self):
        """
        Base data getter for an organization report.
        """
        kwargs = dict(
            organization=self.entity,
            service=self.service,
            period_start=self.period_start,
            generate=self.generate
        )
        if self.service.is_expiring:
            organizations_report = OrganizationReportsService.get_expiring_system_reports(**kwargs)
        else:
            organizations_report = OrganizationReportsService.get_regular_system_reports(**kwargs)

        return organizations_report


class PartnerReportsExportBase(PartnerReportDataMixin, ReportsExportServiceBase):
    """
    This class is a base class for exporting partner reports. It inherits from ReportsExportServiceBase.
    It implements the get_report_data method which fetches the report data for a partner.
    """
    pass


class OrganizationReportsExportBase(OrganizationReportDataMixin, ReportsExportServiceBase):
    """
    This class is a base class for exporting organization reports. It inherits from ReportsExportServiceBase.
    It implements the get_report_data method which fetches the report data for an organization.
    """


class SummaryRowBase:
    """
    Base class for summary rows in a report. This class should not be instantiated directly.
    Instead, use one of the subclasses that implements the required methods.
    """

    def __init__(self,
                 usage: ChannelPartnerUsageReportRecord | OrganizationUsageReportRecord,
                 row_num: int,
                 report_format: ReportFormat = ReportFormat.xlsx):
        self.usage = usage
        self.row_num = row_num
        self.report_format = report_format

    @property
    def used_by(self):
        """
        'Used by' column for services summary report sheets.
        """
        raise NotImplemented("This method must be implemented in the subclass.")

    @property
    def total_price(self):
        """
        Formulae for the total price column.
        """
        if self.report_format == ReportFormat.xlsx:
            return f'=E{self.row_num}*F{self.row_num}+G{self.row_num}*H{self.row_num}'
        else:
            return 0

    def values_list(self) -> List[Any]:
        """
        List of values for the current row.
        """
        return [
            self.usage['service_name'],
            self.used_by,
            self.usage['channels'],
            self.usage['monthly_rate'],
            0,
            self.usage['daily_rate'],
            0,
            self.total_price,
        ]

    def number_formats(self) -> List[Optional[str]]:
        """
        List of number formats for the current row.
        """
        return [
            None,
            None,
            Styling.format_number,
            Styling.format_number,
            Styling.format_currency,
            Styling.format_number,
            Styling.format_currency,
            Styling.format_currency,
        ]

    def font(self) -> List[Optional[Font]]:
        """
        List of fonts for the current row.
        """
        return [
            Styling.font,
            Styling.font,
            Styling.font,
            Styling.font,
            Styling.font,
            Styling.font,
            Styling.font,
            Styling.font_bold,
        ]

    def borders(self) -> List[Optional[Border]]:
        """
        List of borders for the current row.
        """
        return [
            Border(left=Styling.border_medium, right=Styling.border_thin, bottom=Styling.border_thin),
            Border(left=Styling.border_thin, right=Styling.border_thin, bottom=Styling.border_thin),
            Border(left=Styling.border_thin, right=Styling.border_thin, bottom=Styling.border_thin),
            Border(left=Styling.border_thin, right=Styling.border_thin, bottom=Styling.border_thin),
            Border(left=Styling.border_thin, right=Styling.border_thin, bottom=Styling.border_thin),
            Border(left=Styling.border_thin, right=Styling.border_thin, bottom=Styling.border_thin),
            Border(left=Styling.border_thin, right=Styling.border_thin, bottom=Styling.border_thin),
            Border(left=Styling.border_thin, right=Styling.border_medium, bottom=Styling.border_thin),
        ]
    
    def colors(self) -> List[Optional[PatternFill]]:
        """
        List of color fillings for the current row.
        """
        return [
            None,
            None,
            None,
            None,
            Styling.fill_formulae,
            None,
            Styling.fill_formulae,
            None,
        ]

    def zip_values(self):
        """
        Return zipped values and styles for the current row.
        """
        return zip(
            self.values_list(),
            self.number_formats(),
            self.font(),
            self.borders(),
            self.colors()
        )


class ChannelPartnerSummaryRow(SummaryRowBase):
    """
    Summary row for channel partners.
    """

    @property
    def used_by(self):
        """
        'Used by' column for services summary report sheets.
        """
        used_by = []
        if partners := self.usage['used_by_channel_partners']:
            used_by.append(f"{partners} CP")
        if orgs := self.usage['used_by_organizations']:
            used_by.append(f"{orgs} ORG")
        return ', '.join(used_by)


class OrganizationSummaryRow(SummaryRowBase):
    """
    Summary row for organizations.
    """

    @property
    def used_by(self):
        """
        'Used by' column for services summary report sheets.
        """
        return f'{self.usage["used_by"]} SYS'


class RegularServiceChangesRow:
    """
    Generate values for row for regular services report sheets.
    """

    def __init__(self,
                 usage: RegularUsageDetailRecord,
                 row_num: int,
                 report_format: ReportFormat = ReportFormat.xlsx,
                 used_by: str = ''):
        self.usage = usage
        self.row_num = row_num
        self.report_format = report_format
        self.last = self.usage['date'] == TotalUsageDate
        self.used_by = used_by

    @property
    def changed(self):
        """
        'Changed' column for regular services report sheets.
        Period start marked as 'Previous periods'.
        Period end is empty string.
        """
        if self.usage['date'] == BeginningOfPeriodDate:
            return 'Previous periods'

        if self.usage['date'] == TotalUsageDate:
            return ''
        return parser.parse(self.usage['date'])

    def values_list(self) -> List[Any]:
        """
        Values list for the current row.
        """
        return [
            self.used_by,
            self.changed,
            self.usage['channels'],
            self.usage['monthly_rate'],
            self.usage['daily_rate'],
        ]

    def number_formats(self) -> List[Optional[str]]:
        """
        Number formats for the current row.
        """
        return [
            None,
            Styling.format_date if isinstance(self.changed, datetime.date) else None,
            Styling.format_number,
            Styling.format_number,
            Styling.format_number,
        ]

    def font(self) -> List[Optional[Font]]:
        """
        Fonts for the current row.
        """
        if not self.last:
            return [
                Styling.font,
                Styling.font,
                Styling.font,
                Styling.font,
                Styling.font,
            ]
        return [
            Styling.font_italic,
            Styling.font_italic,
            Styling.font_italic,
            Styling.font_italic,
            Styling.font_italic,
        ]

    def borders(self) -> List[Optional[Border]]:
        """
        Borders for the current row.
        """
        if not self.last:
            return [
                Border(left=Styling.border_medium, right=Styling.border_thin),
                Border(left=Styling.border_thin, right=Styling.border_thin),
                Border(left=Styling.border_thin, right=Styling.border_thin),
                Border(left=Styling.border_thin, right=Styling.border_thin),
                Border(left=Styling.border_thin, right=Styling.border_medium),
            ]
        return [
            Border(left=Styling.border_medium, right=Styling.border_thin,
                   bottom=Styling.border_thin, top=Styling.border_dotted),
            Border(left=Styling.border_thin, right=Styling.border_thin,
                   bottom=Styling.border_thin, top=Styling.border_dotted),
            Border(left=Styling.border_thin, right=Styling.border_thin,
                   bottom=Styling.border_thin, top=Styling.border_dotted),
            Border(left=Styling.border_thin, right=Styling.border_thin,
                   bottom=Styling.border_thin, top=Styling.border_dotted),
            Border(left=Styling.border_thin, right=Styling.border_medium,
                   bottom=Styling.border_thin, top=Styling.border_dotted),
        ]

    def zip_values(self):
        """
        Return zipped values and styles for the current row.
        """
        return zip(
            self.values_list(),
            self.number_formats(),
            self.font(),
            self.borders(),
        )


class ExpiringServiceChangesRow:
    """
    Generate values for row for expiring services report sheets.
    """

    def __init__(self,
                 usage: ExpiringUsageDetailRecord,
                 row_num: int,
                 report_format: ReportFormat = ReportFormat.xlsx,
                 used_by: str = ''):
        self.usage = usage
        self.row_num = row_num
        self.report_format = report_format
        # If the expiration date is TotalUsageDate this is the last row in the report.
        self.last = self.usage['expiration_date'] == TotalUsageDate
        self.used_by = used_by

    @property
    def expiration_date(self):
        """
        'Expiration date' column for expiring services report sheets.
        """
        if self.usage['expiration_date'] == BeginningOfPeriodDate:
            return 'Previous periods'

        if self.usage['expiration_date'] == TotalUsageDate:
            return ''
        return parser.parse(self.usage['expiration_date'])

    def values_list(self) -> List[Any]:
        """
        Values list for the current row.
        """
        return [
            self.used_by,
            self.expiration_date,
            self.usage['channels'],
        ]

    def number_formats(self) -> List[Optional[str]]:
        """
        Number formats for the current row.
        """
        return [
            None,
            Styling.format_date if isinstance(self.expiration_date, datetime.date) else None,
            Styling.format_number,

        ]

    def font(self) -> List[Optional[Font]]:
        """
        Fonts for the current row.
        """
        if not self.last:
            return [
                Styling.font,
                Styling.font,
                Styling.font,

            ]
        return [
            Styling.font_italic,
            Styling.font_italic,
            Styling.font_italic,
        ]

    def borders(self) -> List[Optional[Border]]:
        """
        Borders for the current row.
        """
        if not self.last:
            return [
                Border(left=Styling.border_medium, right=Styling.border_thin),
                Border(left=Styling.border_thin, right=Styling.border_thin),
                Border(left=Styling.border_thin, right=Styling.border_medium),
            ]
        return [
            Border(left=Styling.border_medium, right=Styling.border_thin, bottom=Styling.border_thin, top=Styling.border_dotted),
            Border(left=Styling.border_thin, right=Styling.border_thin, bottom=Styling.border_thin, top=Styling.border_dotted),
            Border(left=Styling.border_thin, right=Styling.border_medium, bottom=Styling.border_thin, top=Styling.border_dotted),
        ]

    def zip_values(self):
        return zip(
            self.values_list(),
            self.number_formats(),
            self.font(),
            self.borders(),
        )


class SummarySheetBase(PartnerReportsExportBase):
    """
    Base class for summary sheets.
    """
    columns = 8

    def __init__(
            self,
            report_date: datetime.date,
            period_start: datetime.date,
            entity: ChannelPartner | Organization,
            sheet: Worksheet
    ):
        super().__init__(
            report_date=report_date,
            period_start=period_start,
            entity=entity,
            service=None,
            sheet=sheet
        )
        self.sheet.title = "Summary"

    def get_report_data(self) -> ChannelPartnerUsageReport | OrganizationUsageReportRecord:
        raise NotImplemented("This method must be implemented in the subclass.")

    def fill_services(self):
        for usage_record in self.get_report_data():
            usage_row = self.record_row_class(usage_record, self.current_row)
            for col_idx, (val, fmt, font, border, fill) in enumerate(usage_row.zip_values()):
                cell = self.sheet.cell(row=self.current_row, column=col_idx + self.start_col)
                cell.value = val
                cell.font = font
                if fmt:
                    cell.number_format = fmt
                cell.border = border
                if fill:
                    cell.fill = fill
            self.current_row += 1
        return self.current_row

    def make_footer(self):
        for i in range(self.start_col, self.start_col + self.columns):
            cell = self.sheet.cell(row=self.current_row, column=i)
            left = None
            right = None
            if i == self.start_col:
                left = Styling.border_medium
            if i == self.start_col + self.columns - 1:
                right = Styling.border_medium
            cell.border = Border(top=Styling.border_double, bottom=Styling.border_medium, left=left, right=right)
        cell = self.sheet.cell(row=self.current_row, column=1 + self.columns)
        first_cell = self.sheet.cell(row=self.start_row, column=1 + self.columns)
        last_cell = self.sheet.cell(row=self.current_row - 1, column=1 + self.columns)
        cell.value = f"=SUM({first_cell.coordinate}:{last_cell.coordinate})"
        cell_number_format = Styling.format_currency
        cell.font = Styling.font_bold

    def generate_report_sheet(self):
        self.make_header()
        self.fill_services()
        self.make_footer()


class PartnerSummaryDataMixin:
    def get_report_data(self) -> ChannelPartnerUsageReport:
        """
        Report data for the channel partner summary sheet.
        """
        kwargs = dict(
            channel_partner=self.entity,
            period_start=self.period_start,
            generate=self.generate
        )
        report = ChannelPartnerReportsService.get_channel_partner_report(**kwargs)
        return report


class ChannelPartnerSummarySheet(PartnerSummaryDataMixin, SummarySheetBase):
    """
    Summary sheet for channel partners.
    """
    columns = 8
    record_row_class = ChannelPartnerSummaryRow
    report: ChannelPartnerUsageReport

    def __init__(
            self,
            report_date: datetime.date,
            period_start: datetime.date,
            entity: ChannelPartner,
            sheet: Worksheet
    ):
        super().__init__(
            report_date=report_date,
            period_start=period_start,
            entity=entity,
            sheet=sheet
        )


class OrganizationSummaryDataMixin:
    def get_report_data(self) -> ChannelPartnerUsageReport:
        """
        Report data for the organization summary sheet.
        """
        kwargs = dict(
            organization=self.entity,
            period_start=self.period_start,
            generate=self.generate
        )
        report = OrganizationReportsService.get_organization_report(**kwargs)
        return report


class OrganizationSummarySheet(OrganizationSummaryDataMixin, SummarySheetBase):
    """
    Summary sheet for organizations.
    """
    columns = 8
    record_row_class = OrganizationSummaryRow
    report: OrganizationUsageReportRecord


class ChannelPartnerRegularServiceSheet(PartnerReportsExportBase):
    """
    Regular services report sheet for channel partners.
    """
    columns = 5
    record_row_class = RegularServiceChangesRow

    def __init__(
            self,
            report_date: datetime.date,
            period_start: datetime.date,
            channel_partner: ChannelPartner,
            service: ChannelPartnerService,
            sheet: Worksheet,
    ):
        super().__init__(
            report_date=report_date,
            period_start=period_start,
            entity=channel_partner,
            service=service,
            sheet=sheet
        )

    def set_footer_values(self):
        """
        Set the formulae of the footer row of the report.
        """
        total_cols = [self.start_col + self.columns + i for i in range(-3, 0)]
        for total_col in total_cols:
            cell = self.sheet.cell(row=self.current_row, column=total_col)
            sub_total_cells = [self.sheet.cell(row=row, column=total_col).coordinate for row in self.sub_total_rows]
            if self.current_row == self.start_row:
                # There were no records
                cell.value = 0
            else:
                cell.value = f'={"+".join(sub_total_cells)}'


class OrganizationRegularServiceSheet(OrganizationReportsExportBase):
    """
    Regular services report sheet for organizations.
    """
    columns = 5
    record_row_class = RegularServiceChangesRow

    def __init__(
            self,
            report_date: datetime.date,
            period_start: datetime.date,
            organization: Organization,
            service: ChannelPartnerService,
            sheet: Worksheet,
    ):
        super().__init__(
            report_date=report_date,
            period_start=period_start,
            entity=organization,
            service=service,
            sheet=sheet
        )

    def set_footer_values(self):
        """
        Set the formulae of the footer row of the report.
        """
        total_cols = [self.start_col + self.columns + i for i in range(-3, 0)]
        for total_col in total_cols:
            cell = self.sheet.cell(row=self.current_row, column=total_col)
            sub_total_cells = [self.sheet.cell(row=row, column=total_col).coordinate for row in self.sub_total_rows]
            if self.current_row == self.start_row:
                # There were no records
                cell.value = 0
            else:
                cell.value = f'={"+".join(sub_total_cells)}'


class ChannelPartnerExpiringServiceSheet(PartnerReportsExportBase):
    """
    Expiring services report sheet for channel partners.
    """
    columns = 3
    record_row_class = ExpiringServiceChangesRow

    def __init__(
            self,
            report_date: datetime.date,
            period_start: datetime.date,
            channel_partner: ChannelPartner,
            service: ChannelPartnerService,
            sheet: Worksheet,
    ):
        super().__init__(
            report_date=report_date,
            period_start=period_start,
            entity=channel_partner,
            service=service,
            sheet=sheet
        )

    def set_footer_values(self):
        """
        Set the formulae of the footer row of the report.
        """
        total_col = self.start_col + self.columns - 1
        cell = self.sheet.cell(row=self.current_row, column=total_col)
        sub_total_cells = [self.sheet.cell(row=row, column=total_col).coordinate for row in self.sub_total_rows]
        if self.current_row == self.start_row:
            # There were no records
            cell.value = 0
        else:
            cell.value = f'={"+".join(sub_total_cells)}'


class OrganizationExpiringServiceSheet(OrganizationReportsExportBase):
    """
    Expiring services report sheet for organizations.
    """
    columns = 3
    record_row_class = ExpiringServiceChangesRow

    def __init__(
            self,
            report_date: datetime.date,
            period_start: datetime.date,
            organization: Organization,
            service: ChannelPartnerService,
            sheet: Worksheet,
    ):
        super().__init__(
            report_date=report_date,
            period_start=period_start,
            entity=organization,
            service=service,
            sheet=sheet
        )

    def set_footer_values(self):
        """
        Set the formulae of the footer row of the report.
        """
        total_col = self.start_col + self.columns - 1
        cell = self.sheet.cell(row=self.current_row, column=total_col)
        sub_total_cells = [self.sheet.cell(row=row, column=total_col).coordinate for row in self.sub_total_rows]
        if self.current_row == self.start_row:
            # There were no records
            cell.value = 0
        else:
            cell.value = f'={"+".join(sub_total_cells)}'


class ChannelPartnerReportGenerator:
    """
    Channel partner report generator.
    """
    def __init__(self,
                 channel_partner: ChannelPartner,
                 report_date: datetime.date = None,
                 period_start: datetime.date = None,
                 report_format: ReportFormat = ReportFormat.xlsx):
        self.channel_partner = channel_partner
        self.report_date = report_date if report_date else datetime.date.today()
        self.period_start = period_start or self.report_date.replace(day=1)
        self.wb: Optional[Workbook] = None
        self.archive = None
        self.buf = None
        self.report_format = report_format

    def generate_summary_sheet(self):
        sheet = self.wb['Summary']
        summary_sheet = ChannelPartnerSummarySheet(
            report_date=self.report_date,
            period_start=self.period_start,
            entity=self.channel_partner,
            sheet=sheet
        )
        summary_sheet.generate_report_sheet()

    def generate_regular_services_sheets(self):
        for service in self.channel_partner.services.filter(sub_type=ChannelPartnerService.REGULAR):
            sheet = self.wb.copy_worksheet(self.wb['Regular Service'])
            service_sheet = ChannelPartnerRegularServiceSheet(
                report_date=self.report_date,
                period_start=self.period_start,
                channel_partner=self.channel_partner,
                service=service,
                sheet=sheet
            )
            service_sheet.generate_report_sheet()

    def generate_expiring_services_sheets(self):
        for service in self.channel_partner.services.exclude(sub_type=ChannelPartnerService.REGULAR):
            sheet = self.wb.copy_worksheet(self.wb['Expiring Service'])
            sheet_name = service.name
            idx = 0
            while sheet_name in self.wb.sheetnames:
                idx += 1
                sheet_name = f"{service.name} ({idx})"
            sheet.title = sheet_name

            service_sheet = ChannelPartnerExpiringServiceSheet(
                report_date=self.report_date,
                period_start=self.period_start,
                channel_partner=self.channel_partner,
                service=service,
                sheet=sheet
            )
            service_sheet.generate_report_sheet()

    def generate_report(self):
        template_path = os.path.join(
            settings.BASE_DIR, 'partners/templates/reports/partner_service_usage_report.xlsx')
        self.wb = load_workbook(template_path)
        self.generate_summary_sheet()
        self.generate_regular_services_sheets()
        self.generate_expiring_services_sheets()
        del self.wb['Regular Service']
        del self.wb['Expiring Service']

    def generate_summary_csv_report(self):
        summary_report = ChannelPartnerCSVSummaryReport(
            report_date=self.report_date,
            period_start=self.period_start,
            entity=self.channel_partner,
            service=None
        )
        self.archive.writestr(summary_report.file_name, summary_report.stream().getvalue())

    def generate_regular_services_csv_report(self):
        for service in self.channel_partner.services.filter(sub_type=ChannelPartnerService.REGULAR):
            reg_service_report = ChannelPartnerCSVRegularServiceReport(
                report_date=self.report_date,
                period_start=self.period_start,
                entity=self.channel_partner,
                service=service,
            )
            self.archive.writestr(reg_service_report.file_name, reg_service_report.stream().getvalue())

    def generate_expiring_services_csv_report(self):
        for service in self.channel_partner.services.exclude(sub_type=ChannelPartnerService.REGULAR):
            exp_service_report = ChannelPartnerCSVExpiringServiceReport(
                report_date=self.report_date,
                period_start=self.period_start,
                entity=self.channel_partner,
                service=service,
            )
            self.archive.writestr(exp_service_report.file_name, exp_service_report.stream().getvalue())

    def generate_csv_report(self):
        """
        Generate the report in CSV format.
        """
        self.buf = BytesIO()
        self.archive = zipfile.ZipFile(self.buf, 'w')
        self.generate_summary_csv_report()
        self.generate_regular_services_csv_report()
        self.generate_expiring_services_csv_report()
        self.archive.close()
        return self.buf

    def save_report_file(self):
        """
        For local use only. Do not call in code.
        """
        self.generate_report()
        self.wb.save(f'{self.channel_partner.name} Service Usage Report.xlsx')

    def save_csv_report_file(self):
        """
        For local use only. Do not call in code.
        """
        self.generate_csv_report()
        self.buf.seek(0)
        with open(f'{self.channel_partner.name} Service Usage Report.zip', 'wb') as f:
            f.write(self.buf.read())
        self.buf.close()

    def stream_xlsx(self) -> BytesIO:
        self.generate_report()
        buf = BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf

    def stream_csv(self) -> BytesIO:
        """
        Stream the report in CSV format.
        """
        return self.generate_csv_report()

    def stream(self) -> BytesIO:
        """
        Stream the report in the specified format.
        """
        if self.report_format == ReportFormat.xlsx:
            return self.stream_xlsx()
        elif self.report_format == ReportFormat.csv:
            return self.stream_csv()
        raise ValueError("Unsupported format. Only 'xlsx' is supported.")


class OrganizationReportGenerator:
    """
    Organization report generator.
    """

    def __init__(self,
                 organization: Organization,
                 period_start: datetime.date = None,
                 report_date: datetime.date = None,
                 report_format: ReportFormat = ReportFormat.xlsx):
        self.organization = organization
        self.report_date = report_date or datetime.datetime.now(tz=datetime.timezone.utc).date()
        self.report_format = report_format
        self.period_start = period_start or self.report_date.replace(day=1)
        self.wb = None
        self.archive = None
        self.buf = None

    def generate_summary_sheet(self):
        sheet = self.wb['Summary']
        summary_sheet = OrganizationSummarySheet(
            report_date=self.report_date,
            period_start=self.period_start,
            entity=self.organization,
            sheet=sheet
        )
        summary_sheet.generate_report_sheet()

    def generate_regular_services_sheets(self):
        for service in self.organization.channel_partner.services.filter(sub_type=ChannelPartnerService.REGULAR):
            sheet = self.wb.copy_worksheet(self.wb['Regular Service'])
            service_sheet = OrganizationRegularServiceSheet(
                report_date=self.report_date,
                period_start=self.period_start,
                organization=self.organization,
                service=service,
                sheet=sheet
            )
            service_sheet.generate_report_sheet()

    def generate_expiring_services_sheets(self):
        for service in self.organization.channel_partner.services.exclude(sub_type=ChannelPartnerService.REGULAR):
            sheet = self.wb.copy_worksheet(self.wb['Expiring Service'])
            service_sheet = OrganizationExpiringServiceSheet(
                report_date=self.report_date,
                period_start=self.period_start,
                organization=self.organization,
                service=service,
                sheet=sheet
            )
            service_sheet.generate_report_sheet()

    def generate_report(self):
        template_path = os.path.join(
            settings.BASE_DIR, 'partners/templates/reports/partner_service_usage_report.xlsx')
        self.wb = load_workbook(template_path)
        self.generate_summary_sheet()
        self.generate_regular_services_sheets()
        self.generate_expiring_services_sheets()
        del self.wb['Regular Service']
        del self.wb['Expiring Service']

    def generate_summary_csv_report(self):
        summary_report = OrganizationCSVSummaryReport(
            report_date=self.report_date,
            period_start=self.period_start,
            entity=self.organization,
            service=None
        )
        self.archive.writestr(summary_report.file_name, summary_report.stream().getvalue())

    def generate_regular_services_csv_report(self):
        for service in self.organization.channel_partner.services.filter(sub_type=ChannelPartnerService.REGULAR):
            reg_service_report = OrganizationCSVRegularServiceReport(
                report_date=self.report_date,
                period_start=self.period_start,
                entity=self.organization,
                service=service,
            )
            self.archive.writestr(reg_service_report.file_name, reg_service_report.stream().getvalue())

    def generate_expiring_services_csv_report(self):
        for service in self.organization.channel_partner.services.exclude(sub_type=ChannelPartnerService.REGULAR):
            exp_service_report = OrganizationCSVExpiringServiceReport(
                report_date=self.report_date,
                period_start=self.period_start,
                entity=self.organization,
                service=service,
            )
            self.archive.writestr(exp_service_report.file_name, exp_service_report.stream().getvalue())

    def generate_csv_report(self):
        """
        Generate the report in CSV format.
        """
        self.buf = BytesIO()
        self.archive = zipfile.ZipFile(self.buf, 'w')
        self.generate_summary_csv_report()
        self.generate_regular_services_csv_report()
        self.generate_expiring_services_csv_report()
        self.archive.close()
        return self.buf

    def save_report_file(self):
        """
        For local use only. Do not call in code.
        """
        self.generate_report()
        self.wb.save(f'{self.organization.name} Service Usage Report.xlsx')

    def save_csv_report_file(self):
        """
        For local use only. Do not call in code.
        """
        self.generate_csv_report()
        self.buf.seek(0)
        with open(f'{self.organization.name} Service Usage Report.zip', 'wb') as f:
            f.write(self.buf.read())
        self.buf.close()

    def stream_xlsx(self) -> BytesIO:
        self.generate_report()
        buf = BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf

    def stream_csv(self) -> BytesIO:
        """
        Stream the report in CSV format.
        """
        return self.generate_csv_report()

    def stream(self) -> BytesIO:
        """
        Stream the report in the specified format.
        """
        if self.report_format == ReportFormat.xlsx:
            return self.stream_xlsx()
        elif self.report_format == ReportFormat.csv:
            return self.stream_csv()
        raise ValueError("Unsupported format. Only 'xlsx' is supported.")


class CSVReportBase(HeadersMixin):
    """
    Base class for exporting summary reports in CSV format.
    """
    generate = True
    headers = []
    def __init__(
            self,
            report_date: datetime.date,
            period_start: datetime.date,
            entity: ChannelPartner | Organization,
            service: ChannelPartnerService | None,
    ):
        self.report_date = report_date
        self.period_start = period_start
        self.period_end = period_start + relativedelta(months=1) - relativedelta(days=1)
        self.entity = entity
        self.service = service
        self.buf = StringIO()
        self.csv_writer = csv.writer(self.buf, delimiter=',', dialect='excel')

    def generate_report(self):
        raise NotImplemented("This method must be implemented in the subclass.")

    def stream(self) -> StringIO:
        """
        Stream the report in CSV format.
        """
        self.generate_report()
        self.buf.seek(0)
        return self.buf


    def make_header(self):
        """
        Fill headers texts in the first 3 rows of the report.
        """
        self.csv_writer.writerow([self.report_name_header()])
        self.csv_writer.writerow([self.date_range_header()])
        self.csv_writer.writerow([self.gen_date_header()])
        self.csv_writer.writerow(self.headers)


class CSVSummaryReportBase(CSVReportBase):
    """
    Channel partner CSV report.
    """
    headers = [
        'Service Name',
        'Used By',
        'Channels',
        'Monthly Rate',
        'Fractional Usage',
    ]

    @property
    def file_name(self):
        return f"{self.entity.name} Service Usage Report - Summary.csv"

    def get_report_data(self):
        raise NotImplemented("This method must be implemented in the subclass.")

    def fill_services(self):
        raise NotImplemented("This method must be implemented in the subclass.")

    def generate_report(self):
        self.make_header()
        self.fill_services()

    @staticmethod
    def get_used_by(usage):
        """
        'Used by' column for services summary report sheets.
        """
        raise NotImplemented("This method must be implemented in the subclass.")

    @staticmethod
    def get_service_name(usage):
        return usage['service_name']

    def get_service_row(self, usage):
        return [
            self.get_service_name(usage),
            self.get_used_by(usage),
            usage['channels'],
            usage['monthly_rate'],
            usage['daily_rate'],
        ]

    def fill_services(self):
        for usage in self.get_report_data():
            self.csv_writer.writerow(self.get_service_row(usage))


class ChannelPartnerCSVSummaryReport(PartnerSummaryDataMixin, CSVSummaryReportBase):

    @staticmethod
    def get_used_by(usage):
        """
        'Used by' column for services summary report sheets.
        """
        used_by = []
        if partners := usage['used_by_channel_partners']:
            used_by.append(f"{partners} CP")
        if orgs := usage['used_by_organizations']:
            used_by.append(f"{orgs} ORG")
        return ', '.join(used_by)


class OrganizationCSVSummaryReport(OrganizationSummaryDataMixin, CSVSummaryReportBase):

    @staticmethod
    def get_used_by(usage):
        """
        'Used by' column for services summary report sheets.
        """
        return f'{usage["used_by"]} SYS'


class CSVRegularServiceReportBase(CSVReportBase):
    """
    Base class for exporting regular services reports in CSV format.
    """
    headers = [
        'Used by',
        'Changed',
        'Channels',
        'Monthly Rate',
        'Fractional Usage',
    ]

    @property
    def file_name(self):
        return f"{self.entity.name} Service Usage Report - {self.service.name} ({self.service.id}).csv"

    def get_report_data(self):
        raise NotImplemented("This method must be implemented in the subclass.")

    def fill_services(self):
        for usages in self.get_report_data():
            used_by_name = self.get_used_by(usages)
            for usage in usages['report']:
                self.csv_writer.writerow([used_by_name, *self.get_service_row(usage)])

    def get_service_row(self, usage):
        return [
            self.get_changed(usage),
            usage['channels'],
            usage['monthly_rate'],
            usage['daily_rate'],
        ]

    @staticmethod
    def get_used_by(usages: ChannelPartnerUsageReportRecord | OrganizationUsageReportRecord):
        """
        'Used by' column for service usage report sheets.
        """
        used_by_name = (usages.get('channel_partner_name') or usages.get('organization_name')
                        or usages.get('system_name') or usages.get('system_id'))
        return used_by_name

    @staticmethod
    def get_changed(usage):
        """
        'Changed' column for regular services report sheets.
        """
        if usage['date'] == BeginningOfPeriodDate:
            return 'Previous periods'

        if usage['date'] == TotalUsageDate:
            return 'Total usages'
        return parser.parse(usage['date'])

    def generate_report(self):
        self.make_header()
        self.fill_services()


class ChannelPartnerCSVRegularServiceReport(PartnerReportDataMixin, CSVRegularServiceReportBase):
    pass


class OrganizationCSVRegularServiceReport(OrganizationReportDataMixin, CSVRegularServiceReportBase):
    pass


class CSVExpiringServiceReportBase(CSVReportBase):
    """
    Base class for exporting expiring services reports in CSV format.
    """
    headers = [
        'Used by',
        'Expiration Date',
        'Channels',
    ]

    @property
    def file_name(self):
        return f"{self.entity.name} Service Usage Report - {self.service.name} ({self.service.id}).csv"

    def get_report_data(self):
        raise NotImplemented("This method must be implemented in the subclass.")

    def fill_services(self):
        for usages in self.get_report_data():
            used_by_name = self.get_used_by(usages)
            for usage in usages['report']:
                self.csv_writer.writerow([used_by_name, *self.get_service_row(usage)])

    def get_service_row(self, usage):
        return [
            self.get_expiration_date(usage),
            usage['channels'],
        ]

    @staticmethod
    def get_used_by(usages: ChannelPartnerUsageReportRecord | OrganizationUsageReportRecord):
        """
        'Used by' column for service usage report sheets.
        """
        used_by_name = (usages.get('channel_partner_name') or usages.get('organization_name')
                        or usages.get('system_name') or usages.get('system_id'))
        return used_by_name

    @staticmethod
    def get_expiration_date(usage):
        """
        'Expiration date' column for expiring services report sheets.
        """
        if usage['expiration_date'] == BeginningOfPeriodDate:
            return 'Previous periods'

        if usage['expiration_date'] == TotalUsageDate:
            return 'Total'
        return parser.parse(usage['expiration_date'])

    def generate_report(self):
        self.make_header()
        self.fill_services()


class ChannelPartnerCSVExpiringServiceReport(PartnerReportDataMixin, CSVExpiringServiceReportBase):
    pass


class OrganizationCSVExpiringServiceReport(OrganizationReportDataMixin, CSVExpiringServiceReportBase):
    pass